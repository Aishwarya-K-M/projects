from flask import Blueprint, render_template,url_for,redirect, request, flash
from flask_login import login_required,current_user
from openpyxl import load_workbook
from datetime import date


views = Blueprint('views',__name__)
depart = ['cse','aiml','csbs','it','ece']
path = "E:/html_f/attendence_view/website/static/register_attend.xlsx"

@views.route('/')
@login_required
def home():
    return redirect(url_for('views.details'))



@views.route('/details',methods=['GET', 'POST'])
@login_required
def details():
    if request.method == 'POST':
        dept = request.form.get('dept')
        year = request.form.get('year')
        c = request.form.get('class')
        if dept =='select':
            flash('select valid department.', category='error')
        elif year =='select':
            flash('select valid year.', category='error')
        elif c =='select':
            flash('select valid class', category='error')
        else:
            flash(f'{dept.upper()} {year} {c} Data', category='success')
            global wb, namesheet
            namesheet = f"{dept.upper()}_{year}_{c}"
            wb = load_workbook(path)
            wss = wb.sheetnames
            global ws
            if namesheet not in wss:
                ws = wb.create_sheet(title=namesheet)
                wb.save(path)
            ws = wb[namesheet]           
            return redirect(url_for('views.entry'))

    return render_template("details.html", user=current_user)

@views.route('/entry',methods=['GET', 'POST'])
@login_required
def entry():
    if request.method == 'POST' and request.form.get("save", False):
        name = request.form.get('name').title()
        reg = request.form.get('reg')
        dept = request.form.get('dept').upper()
        da = request.form.get('date')
        reason = request.form.get('reason').title()
        isdate = True
        try:
            d = date.fromisoformat(da)
        except:
            isdate = False
        today = date.today()

        if len(name) < 3:
            flash('Enter valid name.', category='error')
        elif not(reg.isnumeric()) or len(reg) != 12:
            flash('Enter valid register no.', category='error')
        elif dept.lower() not in depart:
            flash('Enter valid Department.', category='error')
        elif not(isdate) or d > today or (today-d).days > 60:
            flash('Enter valid date.', category='error')
        elif len(reason) < 3:
            flash('Enter valid reason.', category='error')
        else:
            ws.append([name,reg,dept,da,reason])
            if ws['A1'].value==None:
                ws.delete_rows(1)
            wb.save(path)
            flash('Added Sucessfully!', category='success')
            return redirect(url_for('views.entry'))
    elif request.method == 'POST' and request.form.get("delete", False):
        row = request.form.get('row')
        if int(row)>ws.max_row:
            flash('Enter valid entry number.', category='error')
        elif not(row.isnumeric()):
            flash('Enter valid entry number.', category='error')
        else:
            ws.delete_rows(int(row))
            wb.save(path)
            flash('Deleted sucessfully!', category='success')
            return redirect(url_for('views.entry'))
        
    return render_template("entry.html", sheet = ws,user=current_user)